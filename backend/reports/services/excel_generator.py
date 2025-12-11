"""
Excel Report Generator using openpyxl.

Generates professional Excel reports with multiple sheets, formatting, and charts.
"""

from io import BytesIO
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule


class ExcelGenerator:
    """
    Generate Excel reports from templates and data.
    """

    # Color definitions (hex without #)
    COLORS = {
        "primary": "FB923C",      # Orange
        "primary_light": "FED7AA",
        "success": "22C55E",      # Green
        "success_light": "BBF7D0",
        "warning": "F59E0B",      # Amber
        "warning_light": "FDE68A",
        "danger": "EF4444",       # Red
        "danger_light": "FECACA",
        "info": "3B82F6",         # Blue
        "info_light": "BFDBFE",
        "dark": "1F2937",
        "light": "F3F4F6",
        "white": "FFFFFF",
        "header": "FB923C",       # Same as primary
    }

    def __init__(self):
        """Initialize Excel generator with styles."""
        self.wb = None
        self._setup_styles()

    def _setup_styles(self):
        """Set up named styles for reuse."""
        # Header style
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color=self.COLORS["header"],
                                       end_color=self.COLORS["header"],
                                       fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data styles
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(horizontal="left", vertical="center")

        # Number alignment
        self.number_alignment = Alignment(horizontal="right", vertical="center")

        # Border
        self.thin_border = Border(
            left=Side(style='thin', color=self.COLORS["light"]),
            right=Side(style='thin', color=self.COLORS["light"]),
            top=Side(style='thin', color=self.COLORS["light"]),
            bottom=Side(style='thin', color=self.COLORS["light"])
        )

        # Title style
        self.title_font = Font(bold=True, size=14, color=self.COLORS["dark"])
        self.subtitle_font = Font(size=11, color="666666")

        # Section header style
        self.section_font = Font(bold=True, size=12, color=self.COLORS["primary"])

        # Alternating row fill
        self.alt_row_fill = PatternFill(start_color=self.COLORS["light"],
                                        end_color=self.COLORS["light"],
                                        fill_type="solid")

    async def generate(
        self,
        template,
        data: Dict[str, Any],
        include_charts: bool = True
    ) -> BytesIO:
        """
        Generate Excel report from template and data.

        Args:
            template: Report template instance
            data: Report data dictionary
            include_charts: Whether to include charts

        Returns:
            BytesIO buffer containing the Excel file
        """
        self.wb = Workbook()

        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)

        # Get structure from template
        structure = template.get_excel_structure()

        # Process each sheet
        for sheet_def in structure:
            self._build_sheet(sheet_def, data, include_charts)

        # Add metadata sheet
        self._add_metadata_sheet(template, data)

        # Save to buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _build_sheet(self, sheet_def: Dict, data: Dict, include_charts: bool):
        """Build a worksheet based on definition."""
        sheet_name = sheet_def.get("name", "Sheet")
        sheet_type = sheet_def.get("type")

        ws = self.wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        if sheet_type == "summary":
            self._build_summary_sheet(ws, sheet_def, data)
        elif sheet_type == "data_table":
            self._build_data_table_sheet(ws, sheet_def, data, include_charts)
        elif sheet_type == "dict_table":
            self._build_dict_table_sheet(ws, sheet_def, data)
        elif sheet_type == "shift_comparison":
            self._build_shift_comparison_sheet(ws, sheet_def, data, include_charts)
        elif sheet_type == "list":
            self._build_list_sheet(ws, sheet_def, data)

    def _build_summary_sheet(self, ws, sheet_def: Dict, data: Dict):
        """Build a summary sheet with sections."""
        row = 1

        # Title
        ws.cell(row=row, column=1, value="Report Summary")
        ws.cell(row=row, column=1).font = self.title_font
        row += 1

        # Date info
        ws.cell(row=row, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
        ws.cell(row=row, column=1).font = self.subtitle_font
        row += 2

        sections = sheet_def.get("sections", [])

        for section in sections:
            section_title = section.get("title", "")
            fields = section.get("fields", [])
            data_key = section.get("data_key")

            # Section header
            ws.cell(row=row, column=1, value=section_title)
            ws.cell(row=row, column=1).font = self.section_font
            row += 1

            # Get data source
            section_data = self._get_nested_value(data, data_key) if data_key else data

            # Process fields
            for field_def in fields:
                if isinstance(field_def, tuple):
                    if len(field_def) == 3:
                        label, key, fmt = field_def
                    else:
                        label, key = field_def
                        fmt = None
                else:
                    label = key = field_def
                    fmt = None

                value = self._get_nested_value(section_data, key) if section_data else None

                # Format value
                if fmt == "percentage":
                    value = f"{value}%" if value is not None else "N/A"
                elif fmt == "currency":
                    value = f"${value:,.2f}" if value is not None else "N/A"
                elif fmt == "percentage_change":
                    if value is not None:
                        arrow = "↑" if value > 0 else "↓" if value < 0 else "→"
                        value = f"{arrow} {abs(value)}%"
                elif value is None:
                    value = "N/A"

                # Write row
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=1).font = Font(bold=True, size=10)
                ws.cell(row=row, column=1).fill = PatternFill(start_color=self.COLORS["light"],
                                                              end_color=self.COLORS["light"],
                                                              fill_type="solid")

                ws.cell(row=row, column=2, value=str(value))
                ws.cell(row=row, column=2).font = self.data_font

                row += 1

            row += 1  # Space between sections

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

    def _build_data_table_sheet(self, ws, sheet_def: Dict, data: Dict, include_charts: bool):
        """Build a data table sheet."""
        data_key = sheet_def.get("data_key")
        columns = sheet_def.get("columns", [])

        table_data = self._get_nested_value(data, data_key)
        if not table_data or not isinstance(table_data, list):
            ws.cell(row=1, column=1, value="No data available")
            return

        # Write header row
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.get("label", col_def.get("key")))
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get("width", 15)

        # Write data rows
        for row_idx, row_data in enumerate(table_data, 2):
            for col_idx, col_def in enumerate(columns, 1):
                key = col_def.get("key")
                value = self._get_nested_value(row_data, key)

                # Format value
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                elif isinstance(value, list):
                    value = ", ".join(str(v.get("type", v) if isinstance(v, dict) else v) for v in value[:5])
                elif value is None:
                    value = "-"

                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "-")
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.thin_border

                # Alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = self.alt_row_fill

        # Apply conditional formatting if specified
        cond_format = sheet_def.get("conditional_format")
        if cond_format:
            self._apply_conditional_formatting(ws, cond_format, columns, len(table_data) + 1)

        # Add chart if requested
        if include_charts and sheet_def.get("include_chart"):
            self._add_chart_to_sheet(ws, len(table_data) + 1, columns, table_data)

    def _build_dict_table_sheet(self, ws, sheet_def: Dict, data: Dict):
        """Build a table from a dictionary."""
        data_key = sheet_def.get("data_key")
        columns = sheet_def.get("columns", [])

        dict_data = self._get_nested_value(data, data_key)
        if not dict_data or not isinstance(dict_data, dict):
            ws.cell(row=1, column=1, value="No data available")
            return

        # Write header row
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.get("label", col_def.get("key")))
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get("width", 20)

        # Write data rows
        row_idx = 2
        for key, value in dict_data.items():
            ws.cell(row=row_idx, column=1, value=str(key))
            ws.cell(row=row_idx, column=1).font = self.data_font
            ws.cell(row=row_idx, column=1).border = self.thin_border

            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=2).font = self.data_font
            ws.cell(row=row_idx, column=2).border = self.thin_border
            ws.cell(row=row_idx, column=2).alignment = self.number_alignment

            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=1).fill = self.alt_row_fill
                ws.cell(row=row_idx, column=2).fill = self.alt_row_fill

            row_idx += 1

    def _build_shift_comparison_sheet(self, ws, sheet_def: Dict, data: Dict, include_charts: bool):
        """Build shift comparison table."""
        data_key = sheet_def.get("data_key")
        columns = sheet_def.get("columns", [])

        shift_data = self._get_nested_value(data, data_key)

        if not shift_data:
            ws.cell(row=1, column=1, value="No shift data available")
            return

        # Convert dict to list if needed
        if isinstance(shift_data, dict):
            table_data = [
                {"shift": shift, **metrics}
                for shift, metrics in shift_data.items()
            ]
        else:
            table_data = shift_data

        # Write header row
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.get("label", col_def.get("key")))
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get("width", 15)

        # Write data rows
        for row_idx, row_data in enumerate(table_data, 2):
            for col_idx, col_def in enumerate(columns, 1):
                key = col_def.get("key")
                value = self._get_nested_value(row_data, key)

                if value is None:
                    value = "-"
                elif key == "shift":
                    value = str(value).title()

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.thin_border

                if row_idx % 2 == 0:
                    cell.fill = self.alt_row_fill

        # Add chart
        if include_charts and len(table_data) > 0:
            self._add_shift_chart(ws, len(table_data) + 1, table_data)

    def _build_list_sheet(self, ws, sheet_def: Dict, data: Dict):
        """Build a simple list sheet."""
        title = sheet_def.get("title", "List")
        data_key = sheet_def.get("data_key")

        list_data = self._get_nested_value(data, data_key)

        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = self.section_font

        if not list_data:
            ws.cell(row=3, column=1, value="No items")
            return

        for idx, item in enumerate(list_data, 3):
            ws.cell(row=idx, column=1, value=f"• {item}")
            ws.cell(row=idx, column=1).font = self.data_font

        ws.column_dimensions['A'].width = 50

    def _add_metadata_sheet(self, template, data: Dict):
        """Add a metadata sheet with report info."""
        ws = self.wb.create_sheet(title="Report Info")

        metadata = [
            ("Report Type", template.report_name),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Mine/Scope", data.get("mine_name", "N/A")),
            ("Date Range", data.get("date_range", "N/A")),
            ("System", "Mine Safety PPE Detection System"),
        ]

        for row_idx, (label, value) in enumerate(metadata, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=str(value))

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _apply_conditional_formatting(self, ws, cond_format: Dict, columns: List, last_row: int):
        """Apply conditional formatting to a column."""
        column_name = cond_format.get("column")
        rules = cond_format.get("rules", [])

        # Find column index
        col_idx = None
        for idx, col in enumerate(columns, 1):
            if col.get("key") == column_name:
                col_idx = idx
                break

        if not col_idx:
            return

        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{last_row}"

        # Apply color scale for numeric columns
        if any("< " in str(r.get("condition", "")) or "> " in str(r.get("condition", "")) for r in rules):
            rule = ColorScaleRule(
                start_type='min',
                start_color='22C55E',  # Green
                mid_type='percentile',
                mid_value=50,
                mid_color='F59E0B',    # Amber
                end_type='max',
                end_color='EF4444'     # Red
            )
            ws.conditional_formatting.add(cell_range, rule)

    def _add_chart_to_sheet(self, ws, data_end_row: int, columns: List, table_data: List):
        """Add a bar chart to a data table sheet."""
        if len(table_data) < 2:
            return

        # Find numeric column for chart
        value_col = None
        label_col = 1

        for idx, col in enumerate(columns, 1):
            key = col.get("key", "")
            if "rate" in key.lower() or "count" in key.lower() or "total" in key.lower():
                value_col = idx
                break

        if not value_col:
            value_col = 2  # Default to second column

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Data Visualization"

        data = Reference(ws, min_col=value_col, min_row=1, max_row=data_end_row)
        labels = Reference(ws, min_col=label_col, min_row=2, max_row=data_end_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        chart.width = 15
        chart.height = 8

        ws.add_chart(chart, f"A{data_end_row + 3}")

    def _add_shift_chart(self, ws, data_end_row: int, table_data: List):
        """Add a chart for shift comparison."""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Shift Performance"
        chart.y_axis.title = "Compliance %"

        # Assuming compliance_rate is in column 4
        data = Reference(ws, min_col=4, min_row=1, max_row=data_end_row)
        labels = Reference(ws, min_col=1, min_row=2, max_row=data_end_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        chart.width = 12
        chart.height = 7

        ws.add_chart(chart, f"A{data_end_row + 3}")

    def _get_nested_value(self, data: Dict, key: str):
        """Get a nested value from a dictionary using dot notation."""
        if not key or not data:
            return None

        keys = key.split(".")
        value = data

        for k in keys:
            if isinstance(value, dict):
                value = value.get(k)
            else:
                return None

        return value
